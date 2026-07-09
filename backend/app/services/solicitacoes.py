"""Serviço de Solicitações (US2 parceiro / US5 gestor). contracts/api.md, research D9.

Busca/filtra/pagina/agrupa sobre o dataset VÁLIDO escopado (R-001). Agrupamento por médico
é apresentacional: ordena por médico antes de paginar e nunca corta um grupo entre páginas
(RF-009) — a página efetiva pode trazer >20 itens.
"""

from collections.abc import Callable
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.filtros.engine import FiltroAplicado
from app.domain.filtros.engine import aplica as aplica_filtros
from app.domain.models import AppUser, Solicitacao
from app.domain.scope import filtra_por_escopo, is_gestor
from app.domain.status import casa_busca_status
from app.services.cores import cor_para
from app.services.dataset import Dataset
from app.services.serialize import money_str, serializa_medico, serializa_solicitacao

LIMIT_PADRAO = 20

# Ordenação: coluna + direção. Padrão = data do pedido, mais recente primeiro (feature 008;
# antes a ordem era sempre por médico). O agrupamento visual por médico (RF-009) só vale
# quando a ordem é pela coluna "cliente" — aí as linhas do mesmo médico ficam contíguas.
SORT_PADRAO = "data_pedido"
DIR_PADRAO = "desc"

# Chave de ordenação por coluna. `None` vira um piso do mesmo tipo p/ não misturar tipos no
# sort. "cliente" ordena por (nome, grupo do médico, código) — mantém o médico contíguo.
_SORT_GETTERS: dict[str, Callable[[Solicitacao], object]] = {
    "codigo": lambda s: s.codigo,
    "cliente": lambda s: ((s.cliente or "").lower(), s.medico_grupo_id or "", s.codigo),
    "data_pedido": lambda s: s.data_pedido,
    "valor": lambda s: s.valor,
    "data_vencimento": lambda s: s.data_vencimento,
    "status": lambda s: s.status,
    "recebido_cliente": lambda s: s.recebido_cliente or Decimal(0),
    "iof": lambda s: s.iof or Decimal(0),
    "taxa_juros_mes": lambda s: s.taxa_juros_mes or Decimal(0),
    "prazo_dias": lambda s: s.prazo_dias if s.prazo_dias is not None else -1,
    "unidade": lambda s: (s.unidade or "").lower(),
    "cashback": lambda s: s.cashback,
    "contratante": lambda s: (s.contratante or "").lower(),
}


def _aplica_escopo_e_filtros(
    dataset: Dataset,
    user: AppUser,
    q: str | None,
    filtros: list[FiltroAplicado] | None,
) -> list[Solicitacao]:
    """Escopo R-001 (sempre 1º) + filtros dinâmicos + busca. Não ordena (ver `_ordena`)."""
    # Escopo primeiro e separado: o filtro de UI nunca amplia o escopo do parceiro.
    itens = filtra_por_escopo(dataset.validas, user)
    itens = aplica_filtros(itens, filtros or [])

    if q:
        termo = q.strip().lower()
        itens = [s for s in itens if _casa_busca(s, termo)]
    return itens


def _ordena(itens: list[Solicitacao], sort: str, direcao: str) -> list[Solicitacao]:
    """Ordena pela coluna `sort` (asc/desc). Coluna/direção inválida cai no padrão."""
    getter = _SORT_GETTERS.get(sort, _SORT_GETTERS[SORT_PADRAO])
    reverso = (direcao if direcao in ("asc", "desc") else DIR_PADRAO) == "desc"
    # `codigo` como desempate estável e determinístico entre linhas de igual valor.
    return sorted(itens, key=lambda s: (getter(s), s.codigo), reverse=reverso)


def _casa_busca(s: Solicitacao, termo: str) -> bool:
    """Busca por código, cliente ou status. O status casa os rótulos EXIBIDOS na UI
    ("Vencido"/"A Vencer") além da chave interna — o usuário busca o que vê."""
    return (
        termo in s.codigo.lower()
        or termo in s.cliente.lower()
        or casa_busca_status(s.status, termo)
    )


def _pagina_sem_cortar_grupo(
    itens: list[Solicitacao], offset: int, limit: int
) -> tuple[list[Solicitacao], bool]:
    """Fatia [offset, offset+limit] e estende até fechar o grupo do médico (RF-009)."""
    total = len(itens)
    if offset >= total:
        return [], False
    fim = min(offset + limit, total)
    # Estende enquanto o próximo item pertence ao mesmo médico do último incluído.
    while fim < total and itens[fim].medico_grupo_id == itens[fim - 1].medico_grupo_id:
        fim += 1
    return itens[offset:fim], fim < total


def _serializa(s: Solicitacao, gestor: bool) -> dict:
    item = serializa_solicitacao(s, incluir_gestor=gestor)
    if gestor:
        item["cor_parceiro"] = cor_para(s.contratante)
    return item


def listar_solicitacoes(
    dataset: Dataset,
    user: AppUser,
    q: str | None = None,
    filtros: list[FiltroAplicado] | None = None,
    limit: int = LIMIT_PADRAO,
    offset: int = 0,
    sort: str = SORT_PADRAO,
    direcao: str = DIR_PADRAO,
) -> dict:
    """Lista paginada/filtrável/ordenável (escopada). Resposta: items, total, has_more."""
    filtradas = _aplica_escopo_e_filtros(dataset, user, q, filtros)
    ordenadas = _ordena(filtradas, sort, direcao)
    gestor = is_gestor(user)
    if sort == "cliente":
        # Só na ordem por médico o agrupamento (RF-009) vale — não corta o grupo entre páginas.
        pagina, has_more = _pagina_sem_cortar_grupo(ordenadas, offset, limit)
    else:
        fim = min(offset + limit, len(ordenadas))
        pagina, has_more = ordenadas[offset:fim], fim < len(ordenadas)
    return {
        "items": [_serializa(s, gestor) for s in pagina],
        "total": len(filtradas),
        "has_more": has_more,
    }


# --- Exportação XLSX (feature 008) -------------------------------------------------------
# Escopada como toda leitura (R-001): passa por `_aplica_escopo_e_filtros`, então o arquivo
# NUNCA carrega linha de outra Contratante nem fora da allowlist de Unidades. As colunas
# gestor-only (Parceiro) só saem na visão do gestor — mesma máscara D5' da serialização JSON.

# Cabeçalho roxo (modelo da planilha-mestre): fundo roxo, texto branco em negrito, centralizado.
# Reutilizado pelos dois exports (Solicitações e lote de Vencimentos).
_HDR_FILL = PatternFill("solid", fgColor="6C4F9E")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")


def _estiliza_cabecalho(ws) -> None:
    """Aplica o estilo roxo na 1ª linha (cabeçalho) da worksheet."""
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN

# Rótulos de status EXIBIDOS na UI (a_pagar→"A Vencer", atrasado→"Vencido") — o arquivo
# espelha o que o usuário vê (frontend/src/lib/format.ts), não a chave interna.
_STATUS_EXIBICAO = {"pago": "Pago", "a_pagar": "A Vencer", "atrasado": "Vencido"}

# Colunas do XLSX: id -> (cabeçalho, getter, tipo de célula). A ordem aqui é a ordem no arquivo.
_EXPORT_COLS: dict[str, tuple[str, Callable[[Solicitacao], object], str]] = {
    "codigo": ("Código", lambda s: s.codigo, "text"),
    "cliente": ("Cliente", lambda s: s.cliente, "text"),
    "contratante": ("Parceiro", lambda s: s.contratante, "text"),
    "data_pedido": ("Data do pedido", lambda s: s.data_pedido, "date"),
    "valor": ("Originação", lambda s: s.valor, "money"),
    "data_vencimento": ("Quitação", lambda s: s.data_vencimento, "date"),
    "status": ("Status", lambda s: _STATUS_EXIBICAO.get(s.status, s.status_label), "text"),
    "recebido_cliente": ("Recebido cliente", lambda s: s.recebido_cliente, "money"),
    "iof": ("IOF", lambda s: s.iof, "money"),
    "taxa_juros_mes": ("Taxa ao mês (%)", lambda s: s.taxa_juros_mes, "percent"),
    "prazo_dias": ("Prazo (dias)", lambda s: s.prazo_dias, "int"),
    "unidade": ("Unidade", lambda s: s.unidade, "text"),
    "cashback": ("Rebate", lambda s: s.cashback, "money"),
}
_EXPORT_GESTOR_ONLY = frozenset({"contratante"})

# Formato de data dos XLSX. As LETRAS de um código de formato são lidas no idioma do programa
# que abre o arquivo (alemão usa T/M/J, francês J/M/A...), então "DD/MM/YYYY" saía como o texto
# literal "DD/06/YYYY" fora de um Excel em inglês. O prefixo LCID `[$-416]` (pt-BR) fixa a
# leitura — é o que o próprio Excel grava numa data brasileira. Não usar código sem o prefixo.
_FMT_DATA = "[$-416]dd/mm/yyyy"

_EXPORT_FMT = {"money": "#,##0.00", "date": _FMT_DATA, "percent": '0.00"%"', "int": "0"}


def _colunas_export(colunas: list[str] | None, gestor: bool) -> list[str]:
    """Colunas a exportar, na ordem canônica. Filtra gestor-only p/ parceiro; ids inválidos
    são ignorados. Sem escolha (ou escolha vazia após filtrar) => todas as permitidas."""
    permitidas = [c for c in _EXPORT_COLS if gestor or c not in _EXPORT_GESTOR_ONLY]
    if not colunas:
        return permitidas
    pedidas = set(colunas)
    return [c for c in permitidas if c in pedidas] or permitidas


def _valor_celula(v: object, tipo: str) -> object:
    """Converte o valor de domínio p/ o que o openpyxl grava nativo (número/data/texto)."""
    if v is None:
        return None
    if tipo in ("money", "brl", "percent"):
        return float(v)  # Decimal -> float p/ o Excel tratar como número
    return v


def exporta_solicitacoes_xlsx(
    dataset: Dataset,
    user: AppUser,
    q: str | None = None,
    filtros: list[FiltroAplicado] | None = None,
    colunas: list[str] | None = None,
    sort: str = SORT_PADRAO,
    direcao: str = DIR_PADRAO,
) -> bytes:
    """Gera o XLSX das solicitações escopadas/filtradas/ordenadas. Retorna os bytes do arquivo."""
    filtradas = _aplica_escopo_e_filtros(dataset, user, q, filtros)
    ordenadas = _ordena(filtradas, sort, direcao)
    cols = _colunas_export(colunas, is_gestor(user))

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitações"
    ws.append([_EXPORT_COLS[c][0] for c in cols])
    _estiliza_cabecalho(ws)
    for s in ordenadas:
        ws.append([_valor_celula(_EXPORT_COLS[c][1](s), _EXPORT_COLS[c][2]) for c in cols])

    # Formato numérico/data por coluna (pula o cabeçalho) + largura confortável.
    for i, c in enumerate(cols, start=1):
        letra = get_column_letter(i)
        fmt = _EXPORT_FMT.get(_EXPORT_COLS[c][2])
        if fmt:
            for cell in ws[letra][1:]:
                cell.number_format = fmt
        ws.column_dimensions[letra].width = 18
    ws.freeze_panes = "A2"  # trava o cabeçalho ao rolar

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Exportação XLSX por lote (aba Vencimentos) ------------------------------------------
# Botão "Exportar" em cada barra de unidade/lote da aba Vencimentos. Layout = modelo da
# planilha-mestre do parceiro (mesmas colunas/nomes/estilo do anexo), SEM "Desconto (-IOF)";
# o código é o gerado pelo portal (feature 009), não o bruto do sheet. Escopado R-001 como
# toda leitura: passa por `filtra_por_escopo`, então nunca carrega linha de outra Contratante
# nem fora da allowlist de Unidades. Endpoint novo de dados → entra na varredura de isolamento.

# Colunas do arquivo: (cabeçalho, getter, tipo). A ordem aqui é a ordem no arquivo.
_LOTE_COLS: list[tuple[str, Callable[[Solicitacao], object], str]] = [
    ("Código", lambda s: s.codigo, "text"),
    ("Cliente", lambda s: s.cliente, "text"),
    ("Originação", lambda s: s.valor, "brl"),
    ("Recebido Cliente", lambda s: s.recebido_cliente, "brl"),
    ("IOF", lambda s: s.iof, "brl"),
    ("Taxa ao Mês", lambda s: s.taxa_juros_mes, "percent"),
    ("Data Pedido", lambda s: s.data_pedido, "date"),
    ("Prazo", lambda s: s.prazo_dias, "int"),
    ("Vencimento", lambda s: s.data_vencimento, "date"),
    ("Unidade Referência", lambda s: s.unidade, "text"),
    ("Rebate", lambda s: s.cashback, "brl"),
]
_LOTE_FMT = {"brl": '"R$" #,##0.00', "date": _FMT_DATA, "percent": '0.00"%"', "int": "0"}


def exporta_lote_xlsx(
    dataset: Dataset,
    user: AppUser,
    unidade: str,
    data_vencimento: date | None,
    contratante: str | None = None,
) -> bytes:
    """XLSX de UM lote (unidade + data de vencimento) no modelo da planilha-mestre.

    Escopo R-001 primeiro: só as solicitações do escopo do usuário. Filtra pela `unidade`,
    e — quando informado — pela `data_vencimento` (lote) e `contratante` (o gestor pode ter a
    mesma unidade em Contratantes distintas). Sem `data_vencimento`, exporta a unidade inteira.
    """
    itens = filtra_por_escopo(dataset.validas, user)
    itens = [
        s
        for s in itens
        if s.unidade == unidade
        and (data_vencimento is None or s.data_vencimento == data_vencimento)
        and (contratante is None or s.contratante == contratante)
    ]
    itens.sort(key=lambda s: s.codigo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimentos"
    ws.append([c[0] for c in _LOTE_COLS])
    _estiliza_cabecalho(ws)
    for s in itens:
        ws.append([_valor_celula(getter(s), tipo) for _, getter, tipo in _LOTE_COLS])

    for i, (_, _, tipo) in enumerate(_LOTE_COLS, start=1):
        letra = get_column_letter(i)
        fmt = _LOTE_FMT.get(tipo)
        if fmt:
            for cell in ws[letra][1:]:
                cell.number_format = fmt
        ws.column_dimensions[letra].width = 20
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resumo_medico(itens_escopo: list[Solicitacao], grupo_id: str | None, gestor: bool) -> dict:
    """Agrega as solicitações (já escopadas) do mesmo médico — card do painel lateral.

    Roda só sobre o que já passou por `filtra_por_escopo` (R-001): o total de antecipação
    do médico nunca soma linha de outro parceiro nem fora da allowlist de Unidades.
    """
    grupo = [s for s in itens_escopo if s.medico_grupo_id == grupo_id]
    n = len(grupo)
    valor_total = sum((s.valor for s in grupo), Decimal("0"))
    total_recebido = sum((s.recebido_cliente or Decimal("0") for s in grupo), Decimal("0"))
    total_rebate = sum((s.cashback for s in grupo), Decimal("0"))
    ticket = valor_total / n if n else Decimal("0")
    desde = min((s.data_pedido for s in grupo), default=None)
    resumo = {
        "n_solicitacoes": n,
        "valor_total": money_str(valor_total),
        "total_recebido_cliente": money_str(total_recebido),
        "total_rebate": money_str(total_rebate),
        "ticket_medio": money_str(ticket),
        "n_pagas": sum(1 for s in grupo if s.status == "pago"),
        "n_a_pagar": sum(1 for s in grupo if s.status == "a_pagar"),
        "n_atrasadas": sum(1 for s in grupo if s.status == "atrasado"),
        "unidades": sorted({s.unidade for s in grupo if s.unidade}),
        "desde": desde.isoformat() if desde else None,
    }
    if gestor:
        # Margem MedFlow só na visão do gestor (mesma máscara D5′ do item).
        total_lucro = sum((s.lucro_operacional or Decimal("0") for s in grupo), Decimal("0"))
        resumo["total_lucro_operacional"] = money_str(total_lucro)
    return resumo


def detalhe_solicitacao(dataset: Dataset, user: AppUser, codigo: str) -> dict | None:
    """Detalhe + médico enriquecido (PII) + resumo agregado do médico, escopado.

    Crítico (R-001): passa pelo MESMO ponto de escopo das demais telas —
    `filtra_por_escopo` (Contratante **E** Unidade∈allowlist, feature 003). Nunca devolve a
    solicitação — nem o médico, nem o resumo — fora do escopo do usuário. None se fora do
    escopo/inexistente (não vaza existência: vira 404 no router).
    """
    gestor = is_gestor(user)
    # Escopo R-001 no ponto único: parceiro só enxerga o próprio Contratante E as Unidades
    # da allowlist; gestor recebe tudo. Buscar o código DENTRO do escopo fecha o vazamento
    # (a busca em `dataset.validas` cru ignorava a allowlist e vazava detalhe+PII de Unidade
    # fora do escopo do parceiro).
    no_escopo = filtra_por_escopo(dataset.validas, user)
    for s in no_escopo:
        if s.codigo != codigo:
            continue
        medico = dataset.medico_de(s.cliente)
        # Resumo do médico: só o mesmo contratante (evita fundir homônimos de outro
        # parceiro na visão consolidada do gestor) e sempre dentro do escopo do usuário.
        do_contratante = [x for x in no_escopo if x.contratante == s.contratante]
        return {
            "solicitacao": _serializa(s, gestor),
            "medico": serializa_medico(medico),
            "resumo_medico": _resumo_medico(do_contratante, s.medico_grupo_id, gestor),
        }
    return None
